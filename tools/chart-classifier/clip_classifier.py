"""
clip_labeler.py  —  ESG 圖表分類工具 (CLIP 零樣本, v1.0)

功能：
  - 掃描 data/<year>/<company>/charts/ 下的所有圖片
  - 用 CLIP 進行 5 類分類：bar / line / pie / map / non_chart
  - 複製圖片到 data/charts/<category>/<year>_<company>_<filename>
  - 輸出 Excel：
      Sheet1  = 年份 × 4 種圖表（bar/line/pie/map）總計
      Sheet2+ = 各公司細項（bar/line/pie/map + non_chart）

使用方式：
  python clip_labeler.py [--data_root DATA_ROOT] [--years 2015 2021 ...]

依賴：
  pip install torch torchvision transformers Pillow openpyxl tqdm
"""

from __future__ import annotations

import argparse
import shutil
import sys
import threading
import tkinter as tk
from pathlib import Path
from tkinter import messagebox, ttk
from typing import Optional

import torch
from PIL import Image
from transformers import CLIPModel, CLIPProcessor


# ── 分類設定 ────────────────────────────────────────────────────────────────

CATEGORIES = ["bar", "line", "pie", "map", "non_chart"]

PROMPTS: dict[str, list[str]] = {
    "bar":       ["a bar chart", "a bar graph", "a column chart"],
    "line":      ["a line chart", "a line graph", "a trend chart"],
    "pie":       ["a pie chart", "a donut chart", "a circular chart"],
    "map":       ["a map with numbers", "a geographic map chart", "a choropleth map"],
    "non_chart": ["a table", "a text paragraph", "a photo", "a logo", "a diagram with text"],
}

# 每類取代表性 prompt 的平均 logit 做最終得分
CLIP_MODEL_ID = "openai/clip-vit-base-patch32"


# ── CLIP 分類器 ─────────────────────────────────────────────────────────────

class CLIPClassifier:
    def __init__(self, device: str = "cpu"):
        self.device = device
        self.model = CLIPModel.from_pretrained(CLIP_MODEL_ID).to(device)
        self.processor = CLIPProcessor.from_pretrained(CLIP_MODEL_ID)

        # 預先編碼所有文字提示
        all_texts: list[str] = []
        self._cat_slices: dict[str, slice] = {}
        idx = 0
        for cat in CATEGORIES:
            texts = PROMPTS[cat]
            self._cat_slices[cat] = slice(idx, idx + len(texts))
            all_texts.extend(texts)
            idx += len(texts)

        inputs = self.processor(text=all_texts, return_tensors="pt", padding=True).to(device)
        with torch.no_grad():
            self._text_feats = self.model.get_text_features(**inputs)
            self._text_feats = self._text_feats / self._text_feats.norm(dim=-1, keepdim=True)

    @torch.no_grad()
    def classify(self, image_path: Path) -> str:
        image = Image.open(image_path).convert("RGB")
        inputs = self.processor(images=image, return_tensors="pt").to(self.device)
        img_feat = self.model.get_image_features(**inputs)
        img_feat = img_feat / img_feat.norm(dim=-1, keepdim=True)

        # 計算各類平均相似度
        sims = (img_feat @ self._text_feats.T).squeeze(0)
        scores: dict[str, float] = {}
        for cat in CATEGORIES:
            s = self._cat_slices[cat]
            scores[cat] = sims[s].mean().item()

        return max(scores, key=lambda c: scores[c])


# ── 路徑掃描 ────────────────────────────────────────────────────────────────

IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".bmp", ".webp"}


def collect_images(data_root: Path, years: Optional[list[str]] = None) -> list[tuple[Path, str, str]]:
    """回傳 [(image_path, year, company), ...]"""
    results: list[tuple[Path, str, str]] = []
    year_dirs = sorted(data_root.iterdir()) if data_root.exists() else []
    for yr_dir in year_dirs:
        if not yr_dir.is_dir():
            continue
        if years and yr_dir.name not in years:
            continue
        for co_dir in sorted(yr_dir.iterdir()):
            charts_dir = co_dir / "charts"
            if not charts_dir.is_dir():
                continue
            for f in sorted(charts_dir.iterdir()):
                if f.suffix.lower() in IMAGE_EXTS:
                    results.append((f, yr_dir.name, co_dir.name))
    return results


# ── 核心執行邏輯 ────────────────────────────────────────────────────────────

def run_labeling(
    data_root: Path,
    out_root: Path,
    years: Optional[list[str]],
    device: str,
    progress_cb=None,        # callback(current, total, label)
    pause_event: Optional[threading.Event] = None,
    paused_event: Optional[threading.Event] = None,
    stop_event: Optional[threading.Event] = None,
):
    """主分類流程，可在背景執行緒中呼叫。"""
    images = collect_images(data_root, years)
    total = len(images)
    if total == 0:
        if progress_cb:
            progress_cb(0, 0, "找不到任何圖片，請確認 data/ 路徑與年份。")
        return {}, []

    clf = CLIPClassifier(device=device)

    # 建立輸出目錄
    for cat in CATEGORIES:
        (out_root / cat).mkdir(parents=True, exist_ok=True)

    # 統計 {year: {company: {cat: int}}}
    stats: dict[str, dict[str, dict[str, int]]] = {}
    log_rows: list[tuple[str, str, str, str]] = []  # (year, company, filename, category)

    for i, (img_path, year, company) in enumerate(images):
        if stop_event and stop_event.is_set():
            break

        # 暫停
        if pause_event and pause_event.is_set():
            if paused_event:
                paused_event.set()
            pause_event.wait()  # 等到 pause_event 被 clear
            if paused_event:
                paused_event.clear()

        if stop_event and stop_event.is_set():
            break

        label = clf.classify(img_path)

        # 複製圖片：data/charts/<cat>/<year>_<company>_<filename>
        dest_name = f"{year}_{company}_{img_path.name}"
        dest = out_root / label / dest_name
        shutil.copy2(img_path, dest)

        # 統計
        stats.setdefault(year, {}).setdefault(company, {cat: 0 for cat in CATEGORIES})
        stats[year][company][label] += 1
        log_rows.append((year, company, img_path.name, label))

        if progress_cb:
            progress_cb(i + 1, total, f"{year}/{company}/{img_path.name} → {label}")

    return stats, log_rows


def export_excel(stats: dict, log_rows: list, out_path: Path):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()

    # ── Sheet 1：年份 × 圖表類型 ──
    ws1 = wb.active
    ws1.title = "年份統計"
    chart_cats = ["bar", "line", "pie", "map"]
    headers = ["年份"] + chart_cats + ["合計"]
    ws1.append(headers)
    bold = Font(bold=True)
    for cell in ws1[1]:
        cell.font = bold
        cell.alignment = Alignment(horizontal="center")

    year_totals: dict[str, dict[str, int]] = {}
    for year, companies in stats.items():
        year_totals[year] = {cat: 0 for cat in chart_cats}
        for co_stats in companies.values():
            for cat in chart_cats:
                year_totals[year][cat] += co_stats.get(cat, 0)

    for year in sorted(year_totals):
        row = [year] + [year_totals[year][c] for c in chart_cats]
        row.append(sum(year_totals[year].values()))
        ws1.append(row)

    # ── Sheet 2+：各公司細項 ──
    for year in sorted(stats):
        ws = wb.create_sheet(title=f"{year}")
        headers2 = ["公司"] + chart_cats + ["non_chart", "合計"]
        ws.append(headers2)
        for cell in ws[1]:
            cell.font = bold
            cell.alignment = Alignment(horizontal="center")

        for company in sorted(stats[year]):
            co = stats[year][company]
            row2 = [company] + [co.get(c, 0) for c in chart_cats] + [co.get("non_chart", 0)]
            row2.append(sum(co.values()))
            ws.append(row2)

    wb.save(out_path)


# ── GUI ─────────────────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self, data_root: Path, out_root: Path, years: Optional[list[str]], device: str):
        super().__init__()
        self.data_root = data_root
        self.out_root = out_root
        self.years = years
        self.device = device

        self.title("ESG Chart Classifier — CLIP")
        self.resizable(False, False)
        self._build_ui()

        self.pause_event = threading.Event()
        self.paused_event = threading.Event()
        self.stop_event = threading.Event()
        self._worker: Optional[threading.Thread] = None
        self._stats: dict = {}
        self._log_rows: list = []

    def _build_ui(self):
        PAD = 16
        BG = "#f5f5f7"
        self.configure(bg=BG)

        frame = tk.Frame(self, bg=BG, padx=PAD, pady=PAD)
        frame.pack(fill=tk.BOTH, expand=True)

        tk.Label(frame, text="ESG Chart Classifier", font=("SF Pro Display", 18, "bold"),
                 bg=BG, fg="#1d1d1f").pack(anchor="w")
        tk.Label(frame, text="使用 CLIP 自動分類 ESG 圖表（bar / line / pie / map / non_chart）",
                 font=("SF Pro Text", 11), bg=BG, fg="#86868b").pack(anchor="w", pady=(2, 12))

        # 進度條
        self.progress_var = tk.DoubleVar(value=0)
        self.bar = ttk.Progressbar(frame, variable=self.progress_var, maximum=100, length=460)
        self.bar.pack(fill=tk.X, pady=(0, 6))

        self.status_var = tk.StringVar(value="就緒。按「開始」以執行分類。")
        tk.Label(frame, textvariable=self.status_var, font=("SF Pro Text", 10),
                 bg=BG, fg="#1d1d1f", wraplength=460, justify="left").pack(anchor="w", pady=(0, 12))

        # 按鈕列
        btn_frame = tk.Frame(frame, bg=BG)
        btn_frame.pack(fill=tk.X)

        self.start_btn = tk.Button(btn_frame, text="▶  開始", command=self._start,
                                   width=10, relief=tk.FLAT, bg="#0071e3", fg="white",
                                   activebackground="#0077ed", font=("SF Pro Text", 12))
        self.start_btn.pack(side=tk.LEFT, padx=(0, 8))

        self.pause_sym = tk.StringVar(value="⏸")
        self.pause_btn = tk.Button(btn_frame, textvariable=self.pause_sym,
                                   command=self._toggle_pause,
                                   width=8, relief=tk.FLAT, bg="#e5e5ea", fg="#1d1d1f",
                                   activebackground="#d1d1d6", font=("SF Pro Text", 12),
                                   state=tk.DISABLED)
        self.pause_btn.pack(side=tk.LEFT, padx=(0, 8))

        self.stop_btn = tk.Button(btn_frame, text="■  停止", command=self._stop,
                                  width=8, relief=tk.FLAT, bg="#ff3b30", fg="white",
                                  activebackground="#ff453a", font=("SF Pro Text", 12),
                                  state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT)

    def _start(self):
        if self._worker and self._worker.is_alive():
            return
        self.stop_event.clear()
        self.pause_event.clear()
        self._stats = {}
        self._log_rows = []
        self.progress_var.set(0)
        self.status_var.set("載入 CLIP 模型中…")
        self.start_btn.config(state=tk.DISABLED)
        self.pause_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.NORMAL)

        self._worker = threading.Thread(target=self._worker_fn, daemon=True)
        self._worker.start()

    def _toggle_pause(self):
        if self.pause_event.is_set():
            self.pause_event.clear()
            self.pause_sym.set("⏸")
            self.status_var.set("繼續執行中…")
        else:
            self.pause_event.set()
            self.pause_sym.set("▶")
            self.status_var.set("已暫停，等待當前圖片處理完成…")

    def _stop(self):
        self.stop_event.set()
        # 若正在暫停，需喚醒讓它能偵測到 stop
        if self.pause_event.is_set():
            self.pause_event.clear()
        self.status_var.set("正在停止…")

    def _worker_fn(self):
        def progress_cb(cur, tot, label):
            pct = cur / tot * 100 if tot else 0
            self.after(0, lambda: self.progress_var.set(pct))
            self.after(0, lambda: self.status_var.set(
                f"[{cur}/{tot}] {label}"
            ))

        stats, log_rows = run_labeling(
            self.data_root, self.out_root, self.years, self.device,
            progress_cb=progress_cb,
            pause_event=self.pause_event,
            paused_event=self.paused_event,
            stop_event=self.stop_event,
        )
        self._stats = stats
        self._log_rows = log_rows
        self.after(0, self._on_done)

    def _on_done(self):
        self.pause_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.DISABLED)
        self.start_btn.config(state=tk.NORMAL)

        if self.stop_event.is_set():
            self.status_var.set("已中止。")
            return

        if not self._stats:
            self.status_var.set("完成，但未找到任何圖片。")
            return

        # 匯出 Excel
        total_imgs = len(self._log_rows)
        excel_path = self.out_root / "clip_labeling_results.xlsx"
        export_excel(self._stats, self._log_rows, excel_path)
        self.progress_var.set(100)
        self.status_var.set(
            f"完成！共分類 {total_imgs} 張圖片。結果已存至：\n{excel_path}"
        )
        messagebox.showinfo("完成", f"共分類 {total_imgs} 張圖片。\nExcel 已存至：\n{excel_path}")


# ── 入口 ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="CLIP-based ESG chart labeler")
    parser.add_argument("--data_root", type=str, default=None,
                        help="data/ 根目錄（預設：同 repo 的 data/）")
    parser.add_argument("--years", nargs="+", default=None,
                        help="只處理指定年份，例如 --years 2021 2022")
    parser.add_argument("--no_gui", action="store_true",
                        help="無 GUI 模式（直接執行）")
    args = parser.parse_args()

    # 自動偵測 data_root
    if args.data_root:
        data_root = Path(args.data_root)
    else:
        # 從本檔案往上兩層找 data/
        data_root = Path(__file__).resolve().parent.parent.parent / "data"

    out_root = data_root / "charts"

    # 裝置偵測
    if torch.cuda.is_available():
        device = "cuda"
    elif hasattr(torch.backends, "mps") and torch.backends.mps.is_available():
        device = "mps"
    else:
        device = "cpu"

    print(f"使用裝置：{device}")
    print(f"data_root：{data_root}")
    print(f"out_root ：{out_root}")

    if args.no_gui:
        # 純命令列模式
        stats, log_rows = run_labeling(
            data_root, out_root, args.years, device,
            progress_cb=lambda cur, tot, lbl: print(f"[{cur}/{tot}] {lbl}"),
        )
        if stats:
            excel_path = out_root / "clip_labeling_results.xlsx"
            export_excel(stats, log_rows, excel_path)
            print(f"\n完成！Excel 已存至：{excel_path}")
    else:
        app = App(data_root, out_root, args.years, device)
        app.mainloop()


if __name__ == "__main__":
    main()
