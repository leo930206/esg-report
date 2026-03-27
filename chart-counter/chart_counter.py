#!/usr/bin/env python3
"""
ESG 圖表計數器 v1.1
使用 CLIP zero-shot 分類，統計各公司 ESG 報告中的圖表數量。
判斷標準：含統計數字的圖示（長條、折線、圓餅、散點等）與表格，排除 logo、裝飾圖、照片。
"""

import os
import platform
import subprocess
import threading
import queue
import time
import tkinter as tk
import tkinter.ttk as ttk
from tkinter import scrolledtext, messagebox
from pathlib import Path
from datetime import datetime

__version__ = "1.1"

# ── 路徑 ─────────────────────────────────────────────────────────────
DATA_DIR   = Path(__file__).parent.parent / "data"
OUTPUT_XLS = DATA_DIR / "chart_statistics.xlsx"

# ── CLIP 設定 ─────────────────────────────────────────────────────────
CLIP_MODEL_ID   = "openai/clip-vit-base-patch32"
CHART_THRESHOLD = 0.55
BATCH_SIZE      = 16

CHART_PROMPTS = [
    "a statistical data visualization such as bar chart, line graph, pie chart, "
    "scatter plot, or a data table with numbers and statistics",
    "a non-data image such as company logo, photograph, decorative graphic, "
    "icon, map, or plain text without statistical numbers",
]

# ── Apple 樣式 ────────────────────────────────────────────────────────
APPLE_BG     = '#f5f5f7'
APPLE_CARD   = '#ffffff'
APPLE_TEXT   = '#1d1d1f'
APPLE_GREY   = '#86868b'
APPLE_BLUE   = '#0071e3'
APPLE_GREEN  = '#34c759'
APPLE_RED    = '#ff3b30'
APPLE_ORANGE = '#ff9500'
APPLE_BORDER = '#d2d2d7'
FONT_MAIN    = ('Helvetica Neue', 10)
FONT_LABEL   = ('Helvetica Neue', 9)

# ── 全域狀態 ──────────────────────────────────────────────────────────
log_queue       = queue.Queue()
stop_event      = threading.Event()   # 關閉時通知 worker 退出
done_event      = threading.Event()
pause_requested = threading.Event()   # 使用者按下暫停，worker 跑完當前公司後觸發
resume_event    = threading.Event()   # set=執行中，clear=暫停（worker 在此阻塞）
paused_event    = threading.Event()   # worker 已進入暫停（GUI 用於偵測）
resume_event.set()

ui_stats: dict = {
    'total': 0, 'chart': 0, 'non_chart': 0, 'error': 0,
}

# {year: [(sid, name, chart_count), ...]}
results: dict[str, list[tuple[str, str, int]]] = {}

_clip_model     = None
_clip_processor = None


# ── 工具函式 ──────────────────────────────────────────────────────────
def set_app_icon(root: tk.Tk) -> None:
    """載入 ESG.png 設定 Dock 圖示與 tkinter 視窗圖示。"""
    icon_path = Path(__file__).parent.parent / "ESG.png"
    if not icon_path.exists():
        return
    try:
        from AppKit import NSApplication, NSImage
        ns_img = NSImage.alloc().initWithContentsOfFile_(str(icon_path))
        if ns_img:
            NSApplication.sharedApplication().setApplicationIconImage_(ns_img)
    except Exception:
        pass
    try:
        photo = tk.PhotoImage(file=str(icon_path))
        root.iconphoto(True, photo)
        root._icon_ref = photo
    except Exception:
        pass


def _log(level: str, msg: str) -> None:
    ts = datetime.now().strftime('%H:%M:%S')
    log_queue.put((level, ts, msg))


def _open_folder(path: str) -> None:
    if platform.system() == 'Windows':
        os.startfile(path)
    elif platform.system() == 'Darwin':
        subprocess.Popen(['open', path])
    else:
        subprocess.Popen(['xdg-open', path])


def _get_device() -> str:
    import torch
    if torch.backends.mps.is_available():
        return 'mps'
    if torch.cuda.is_available():
        return 'cuda'
    return 'cpu'


def _company_parts(folder_name: str) -> tuple[str, str]:
    """回傳 (股票代碼, 公司名稱)，解析失敗則 ('', folder_name)。"""
    parts = folder_name.split('_', 2)
    if len(parts) >= 3:
        return parts[1], parts[2]
    return '', folder_name


# ── CLIP 分類邏輯 ─────────────────────────────────────────────────────
def _load_clip(device: str):
    global _clip_model, _clip_processor
    if _clip_model is not None:
        return _clip_model, _clip_processor
    from transformers import CLIPModel, CLIPProcessor
    _log('info', '載入 CLIP 模型中（首次執行需下載約 600 MB）…')
    _clip_model     = CLIPModel.from_pretrained(CLIP_MODEL_ID).to(device)
    _clip_processor = CLIPProcessor.from_pretrained(CLIP_MODEL_ID)
    _clip_model.eval()
    _log('info', f'CLIP 模型載入完成，裝置：{device}')
    return _clip_model, _clip_processor


def _classify_batch(images: list, model, processor, device: str, threshold: float) -> list[bool]:
    import torch
    with torch.no_grad():
        inputs = processor(
            text=CHART_PROMPTS, images=images,
            return_tensors='pt', padding=True,
        )
        inputs = {k: v.to(device) for k, v in inputs.items()}
        logits = model(**inputs).logits_per_image
        probs  = logits.softmax(dim=1).cpu().numpy()
    return [float(p[0]) >= threshold for p in probs]


# ── 分類執行緒 ────────────────────────────────────────────────────────
def run_classification(years: list[str], threshold: float) -> None:
    global results, ui_stats
    results  = {}
    ui_stats = {'total': 0, 'chart': 0, 'non_chart': 0, 'error': 0}
    stop_event.clear()
    done_event.clear()
    pause_requested.clear()
    resume_event.set()
    paused_event.clear()

    try:
        import torch
        from PIL import Image
    except ImportError as e:
        _log('error', f'缺少套件：{e}')
        _log('error', '請執行：pip install transformers torch Pillow')
        done_event.set()
        return

    device = _get_device()
    try:
        model, processor = _load_clip(device)
    except Exception as e:
        _log('error', f'模型載入失敗：{e}')
        done_event.set()
        return

    from PIL import Image

    for year in sorted(years):
        if stop_event.is_set():
            break
        year_dir = DATA_DIR / year
        if not year_dir.is_dir():
            _log('warning', f'找不到資料夾：{year_dir}')
            continue

        results[year] = []
        company_dirs  = sorted(d for d in year_dir.iterdir() if d.is_dir())
        if not company_dirs:
            _log('warning', f'{year}：無公司子資料夾')
            continue

        for company_dir in company_dirs:
            if stop_event.is_set():
                break

            images_dir = company_dir / 'images'
            if not images_dir.is_dir():
                continue
            jpg_files = sorted(images_dir.glob('*.jpg'))
            if not jpg_files:
                continue

            sid, display = _company_parts(company_dir.name)
            chart_count  = 0

            for i in range(0, len(jpg_files), BATCH_SIZE):
                if stop_event.is_set():
                    break
                batch_paths = jpg_files[i:i + BATCH_SIZE]
                images_pil: list = []
                for p in batch_paths:
                    try:
                        images_pil.append(Image.open(p).convert('RGB'))
                        ui_stats['total'] += 1
                    except Exception as e:
                        _log('warning', f'無法讀取：{p.name}（{e}）')
                        ui_stats['error'] += 1

                if not images_pil:
                    continue

                try:
                    flags = _classify_batch(images_pil, model, processor, device, threshold)
                    for is_chart in flags:
                        if is_chart:
                            chart_count       += 1
                            ui_stats['chart'] += 1
                        else:
                            ui_stats['non_chart'] += 1
                except Exception as e:
                    _log('error', f'CLIP 分類失敗（{display} batch {i}）：{e}')
                    ui_stats['error'] += len(images_pil)

            results[year].append((sid, display, chart_count))
            _log('info', f'{year}  {sid} {display}：{chart_count} 張圖表')

            # ── 暫停點：公司處理完後，若有暫停請求則存檔並阻塞 ──
            if pause_requested.is_set():
                pause_requested.clear()
                _save_excel()
                _log('info', '⏸ 已暫停並存檔，可安全關閉視窗')
                paused_event.set()
                resume_event.clear()
                resume_event.wait()   # 阻塞直到繼續或關閉
                paused_event.clear()
                if stop_event.is_set():
                    break

        if year in results:
            _log('info', f'── {year} 完成，共 {sum(c for _, _, c in results[year])} 張圖表 ──')

    if not stop_event.is_set():
        _save_excel()
    else:
        _log('warning', '已中止')

    done_event.set()


# ── Excel 輸出 ────────────────────────────────────────────────────────
def _save_excel() -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        _log('error', '缺少 openpyxl，請執行：pip install openpyxl')
        return

    wb       = openpyxl.Workbook()
    hdr_fill = PatternFill('solid', fgColor='0071E3')
    hdr_font = Font(bold=True, color='FFFFFF', name='Helvetica Neue', size=10)
    ctr      = Alignment(horizontal='center')

    ws1 = wb.active
    ws1.title = '總覽'
    ws1.column_dimensions['A'].width = 10
    ws1.column_dimensions['B'].width = 14
    for col, title in enumerate(['年度', '圖表總數'], 1):
        c = ws1.cell(1, col, title)
        c.font, c.fill, c.alignment = hdr_font, hdr_fill, ctr

    all_years = [str(y) for y in range(2015, 2025)]
    for row_i, year in enumerate(all_years, 2):
        total = sum(c for _, _, c in results.get(year, []))
        ws1.cell(row_i, 1, year).alignment = ctr
        ws1.cell(row_i, 2, total).alignment = ctr

    for year in all_years:
        ws = wb.create_sheet(title=str(year))
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 28
        ws.column_dimensions['C'].width = 12
        for col, title in enumerate(['公司代碼', '公司名稱', '圖表數'], 1):
            c = ws.cell(1, col, title)
            c.font, c.fill, c.alignment = hdr_font, hdr_fill, ctr
        for row_i, (sid, name, count) in enumerate(
            sorted(results.get(year, []), key=lambda x: x[0]), 2
        ):
            ws.cell(row_i, 1, sid).alignment = ctr
            ws.cell(row_i, 2, name)
            ws.cell(row_i, 3, count).alignment = ctr

    OUTPUT_XLS.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLS)
    _log('info', f'Excel 已儲存：{OUTPUT_XLS}')


# ── GUI ───────────────────────────────────────────────────────────────
class App:
    def __init__(self) -> None:
        self.root = tk.Tk()
        self.root.title('ESG 圖表計數器')
        self.root.configure(bg=APPLE_BG)
        self.root.resizable(False, False)
        self.root.protocol('WM_DELETE_WINDOW', self._on_close)
        set_app_icon(self.root)

        self._year_vars: dict[str, tk.BooleanVar] = {}
        self._running  = False
        self._paused   = False

        self._build_header()
        self._build_year_picker()
        self._build_threshold()
        self._build_clip_info()
        self._build_stat_cards()
        self._build_log()
        self._build_footer()

        self.root.after(200, self._poll)

    # ── 各區塊 ────────────────────────────────────────────────────
    def _build_header(self) -> None:
        hdr = tk.Frame(self.root, bg=APPLE_BLUE)
        hdr.pack(fill=tk.X)
        tk.Label(hdr, text='📊  ESG 圖表計數器',
                 font=('Helvetica Neue', 14, 'bold'),
                 fg='white', bg=APPLE_BLUE, pady=12, padx=16).pack(side=tk.LEFT)

        btn_frame = tk.Frame(hdr, bg=APPLE_BLUE)
        btn_frame.pack(side=tk.RIGHT, padx=12)

        self._run_btn = tk.Button(
            btn_frame, text='▶  開始分類',
            font=FONT_MAIN, bg='white', fg=APPLE_BLUE,
            relief='flat', padx=10, pady=4, cursor='hand2',
            command=self._on_run,
        )
        self._run_btn.pack(side=tk.LEFT, padx=(0, 6))

        self._pause_btn = tk.Button(
            btn_frame, text='⏸  暫停',
            font=FONT_MAIN, bg='#ff9f0a', fg='white',
            relief='flat', padx=10, pady=4, cursor='hand2',
            state='disabled', command=self._on_pause_resume,
        )
        self._pause_btn.pack(side=tk.LEFT, padx=(0, 6))

        tk.Button(
            btn_frame, text='📁  開啟輸出',
            font=FONT_MAIN, bg='white', fg=APPLE_TEXT,
            relief='flat', padx=10, pady=4, cursor='hand2',
            command=lambda: _open_folder(str(DATA_DIR)),
        ).pack(side=tk.LEFT)

    def _build_clip_info(self) -> None:
        frame = tk.Frame(self.root, bg='#e8f0fc',
                         highlightthickness=1, highlightbackground='#b8cef7')
        frame.pack(fill=tk.X, padx=16, pady=(10, 0))
        tk.Label(
            frame,
            text='📖 圖表門檻選擇說明',
            font=('Helvetica Neue', 9, 'bold'),
            fg='#1a56c4', bg='#e8f0fc',
        ).pack(anchor='w', padx=10, pady=(6, 0))
        tk.Label(
            frame,
            text=(
                '程式會為每張圖片評分（0～1），只有達到「圖表門檻」分數的圖片才會被計入。\n'
                '  • 分數高（門檻高）→ 標準嚴格，只保留最確定的圖表，logo 或照片較不易混入，且圖表可能遺漏\n'
                '  • 分數低（門檻低）→ 標準寬鬆，圖表抓得多，但非圖表也容易混入\n'
            ),
            font=FONT_LABEL,
            fg='#1d3557', bg='#e8f0fc',
            justify='left',
        ).pack(anchor='w', padx=10, pady=(2, 8))

    def _build_year_picker(self) -> None:
        frame = tk.LabelFrame(self.root, text=' 選擇年度 ',
                              font=FONT_LABEL, bg=APPLE_BG,
                              fg=APPLE_GREY, bd=1, relief='groove')
        frame.pack(fill=tk.X, padx=16, pady=(10, 0))

        inner = tk.Frame(frame, bg=APPLE_BG)
        inner.pack(fill=tk.X, padx=8, pady=6)
        for i, year in enumerate(range(2015, 2025)):
            var = tk.BooleanVar(value=True)
            self._year_vars[str(year)] = var
            tk.Checkbutton(
                inner, text=str(year), variable=var,
                font=FONT_MAIN, bg=APPLE_BG, fg=APPLE_TEXT,
                activebackground=APPLE_BG, selectcolor=APPLE_CARD,
            ).grid(row=0, column=i, padx=4)

        ctrl = tk.Frame(frame, bg=APPLE_BG)
        ctrl.pack(anchor='e', padx=8, pady=(0, 4))
        tk.Button(ctrl, text='全選', font=FONT_LABEL, fg=APPLE_BLUE,
                  bg=APPLE_BG, relief='flat', cursor='hand2',
                  command=lambda: [v.set(True)  for v in self._year_vars.values()]
                  ).pack(side=tk.LEFT, padx=2)
        tk.Button(ctrl, text='全不選', font=FONT_LABEL, fg=APPLE_BLUE,
                  bg=APPLE_BG, relief='flat', cursor='hand2',
                  command=lambda: [v.set(False) for v in self._year_vars.values()]
                  ).pack(side=tk.LEFT, padx=2)

    def _build_threshold(self) -> None:
        frame = tk.Frame(self.root, bg=APPLE_BG)
        frame.pack(fill=tk.X, padx=16, pady=(8, 0))
        tk.Label(frame, text='圖表門檻（CLIP chart 機率 ≥）',
                 font=FONT_LABEL, fg=APPLE_GREY, bg=APPLE_BG).pack(side=tk.LEFT)
        self._threshold_var = tk.DoubleVar(value=CHART_THRESHOLD)
        tk.Scale(frame, variable=self._threshold_var,
                 from_=0.3, to=0.9, resolution=0.05,
                 orient='horizontal', length=180,
                 bg=APPLE_BG, fg=APPLE_TEXT, highlightthickness=0,
                 troughcolor=APPLE_BORDER, sliderrelief='flat',
                 ).pack(side=tk.LEFT, padx=(8, 0))
        tk.Label(frame, textvariable=self._threshold_var,
                 font=FONT_LABEL, fg=APPLE_BLUE, bg=APPLE_BG, width=4).pack(side=tk.LEFT)

    def _build_stat_cards(self) -> None:
        frame = tk.Frame(self.root, bg=APPLE_BG)
        frame.pack(fill=tk.X, padx=16, pady=10)
        self._stat_vars = {
            '總圖片':    tk.StringVar(value='0'),
            '✅ 圖表':   tk.StringVar(value='0'),
            '❌ 非圖表': tk.StringVar(value='0'),
            '⚠️ 錯誤':  tk.StringVar(value='0'),
        }
        colors = [APPLE_TEXT, APPLE_GREEN, APPLE_RED, APPLE_ORANGE]
        for i, (label, var) in enumerate(self._stat_vars.items()):
            card = tk.Frame(frame, bg=APPLE_CARD,
                            highlightthickness=1, highlightbackground=APPLE_BORDER)
            card.grid(row=0, column=i, padx=5, sticky='ew')
            frame.columnconfigure(i, weight=1)
            tk.Label(card, text=label, font=FONT_LABEL,
                     fg=APPLE_GREY, bg=APPLE_CARD).pack(pady=(8, 0))
            tk.Label(card, textvariable=var,
                     font=('Helvetica Neue', 18, 'bold'),
                     fg=colors[i], bg=APPLE_CARD).pack(pady=(2, 8))

    def _build_log(self) -> None:
        frame = tk.Frame(self.root, bg=APPLE_BG)
        frame.pack(fill=tk.BOTH, expand=True, padx=16, pady=(0, 4))
        tk.Label(frame, text='執行紀錄', font=FONT_LABEL,
                 fg=APPLE_GREY, bg=APPLE_BG).pack(anchor='w')
        self._log_text = scrolledtext.ScrolledText(
            frame, width=80, height=16,
            font=('Menlo', 9), bg='#1e1e1e', fg='#d4d4d4',
            insertbackground='white', relief='flat', state='disabled',
        )
        self._log_text.pack(fill=tk.BOTH, expand=True)
        for tag, color in [
            ('info', '#d4d4d4'), ('warning', '#ffc107'),
            ('error', '#ff6b6b'), ('skip', '#86868b'),
        ]:
            self._log_text.tag_configure(tag, foreground=color)

    def _build_footer(self) -> None:
        self._status_var = tk.StringVar(value='就緒')
        tk.Label(self.root, textvariable=self._status_var,
                 font=FONT_LABEL, fg=APPLE_GREY, bg=APPLE_BG,
                 anchor='w').pack(fill=tk.X, padx=16, pady=(0, 8))

    # ── 事件處理 ──────────────────────────────────────────────────
    def _on_run(self) -> None:
        years = [y for y, v in self._year_vars.items() if v.get()]
        if not years:
            messagebox.showwarning('未選年度', '請至少選擇一個年度。')
            return
        if self._running:
            return

        self._running = True
        self._paused  = False
        self._run_btn.config(state='disabled')
        self._pause_btn.config(state='normal', text='⏸  暫停', bg='#ff9f0a')
        self._status_var.set('分類中…')

        self._log_text.config(state='normal')
        self._log_text.delete('1.0', tk.END)
        self._log_text.config(state='disabled')

        threshold = float(self._threshold_var.get())
        threading.Thread(
            target=run_classification,
            args=(years, threshold),
            daemon=True,
        ).start()

    def _on_pause_resume(self) -> None:
        if not self._running:
            return
        if self._paused:
            # 繼續
            self._paused = False
            resume_event.set()
            self._pause_btn.config(text='⏸  暫停', bg='#ff9f0a')
            self._status_var.set('分類中…')
            _log('info', '▶ 繼續執行')
        else:
            # 請求暫停（跑完當前公司後生效）
            pause_requested.set()
            self._pause_btn.config(state='disabled')
            self._status_var.set('等待當前公司完成後暫停…')

    def _on_close(self) -> None:
        if self._running and not self._paused:
            messagebox.showwarning(
                '尚在執行中',
                '請先按「⏸ 暫停」，等目前這間公司處理完後再關閉視窗。',
            )
            return
        stop_event.set()
        resume_event.set()   # 讓 worker 從暫停中醒來並偵測 stop
        self.root.destroy()

    # ── 輪詢更新 ──────────────────────────────────────────────────
    def _poll(self) -> None:
        while True:
            try:
                level, ts, msg = log_queue.get_nowait()
            except queue.Empty:
                break
            self._log_text.config(state='normal')
            self._log_text.insert(tk.END, f'[{ts}] ', 'skip')
            self._log_text.insert(tk.END, msg + '\n', level)
            self._log_text.see(tk.END)
            self._log_text.config(state='disabled')

        labels = ['總圖片', '✅ 圖表', '❌ 非圖表', '⚠️ 錯誤']
        keys   = ['total',  'chart',  'non_chart',  'error']
        for label, key in zip(labels, keys):
            self._stat_vars[label].set(str(ui_stats[key]))

        # worker 已進入暫停狀態
        if paused_event.is_set() and not self._paused:
            self._paused = True
            self._pause_btn.config(state='normal', text='▶  繼續', bg=APPLE_GREEN)
            self._status_var.set('⏸ 已暫停並存檔，可安全關閉視窗')

        if done_event.is_set():
            done_event.clear()
            self._running = False
            self._paused  = False
            self._run_btn.config(state='normal')
            self._pause_btn.config(state='disabled', text='⏸  暫停', bg='#ff9f0a')
            self._status_var.set('✅ 已完成')

        self.root.after(250, self._poll)

    def run(self) -> None:
        self.root.mainloop()


# ── 主程式 ────────────────────────────────────────────────────────────
if __name__ == '__main__':
    App().run()
